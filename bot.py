import os
import logging
import asyncio
import glob
import re
from contextlib import asynccontextmanager

import ijson
import requests
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from fastapi import FastAPI, Request, Response, status
from telegram import Update
from telegram.ext import Application, CommandHandler, MessageHandler, filters, ContextTypes

# Cấu hình log hệ thống
logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s', 
    level=logging.INFO
)
logger = logging.getLogger(__name__)

BOT_TOKEN = os.getenv("BOT_TOKEN")
RENDER_EXTERNAL_URL = os.getenv("RENDER_EXTERNAL_URL") 
PORT = int(os.getenv("PORT", 8000))

tg_application = Application.builder().token(BOT_TOKEN).build()

# --- HÀM LÀM SẠCH HỆ THỐNG ---
def clear_system_cached_files():
    count = 0
    patterns = ["temp_raw_*", "*_converted.xlsx"]
    for pattern in patterns:
        for filepath in glob.glob(pattern):
            try:
                if os.path.exists(filepath):
                    os.remove(filepath)
                    count += 1
            except Exception as e:
                logger.error(f"❌ Lỗi xóa file tạm {filepath}: {str(e)}")
    return count

# --- TẢI FILE TỪ RAW URL THEO TỪNG CHUNK (HỖ TRỢ FILE LỚN, KHÔNG NGẬM RAM) ---
def download_url_stream(url, dest_path):
    logger.info(f"🌐 Đang kết nối Stream tới URL: {url}")
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
    }
    with requests.get(url, headers=headers, stream=True) as r:
        r.raise_for_status()
        with open(dest_path, 'wb') as f:
            for chunk in r.iter_content(chunk_size=65536):  # Stream block 64KB
                f.write(chunk)

# --- GIẢI MÃ SINGLE-PASS TỪ FILE JSON RA EXCEL ---
def convert_heavy_json_to_excel(json_filepath, excel_filepath, queue, loop):
    try:
        logger.info(f"⚡ Bắt đầu bóc tách dữ liệu Single-Pass: {json_filepath}")
        
        wb = openpyxl.Workbook(write_only=True)
        ws = wb.create_sheet(title="DiemThi")
        ws.views.sheetView[0].showGridLines = True
        
        # Styles tiêu chuẩn đẹp (Segoe UI, Xanh đậm)
        header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        data_font = Font(name="Segoe UI", size=10)
        sbd_alignment = Alignment(horizontal="center", vertical="center")
        score_alignment = Alignment(horizontal="right", vertical="center")
        thin_side = Side(border_style="thin", color="D9D9D9")
        border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        
        zebra_fill = PatternFill(start_color="F2F5F8", end_color="F2F5F8", fill_type="solid")
        white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

        cols = []
        header_written = False
        student_count = 0

        with open(json_filepath, 'rb') as f:
            parser = ijson.parse(f)
            current_sbd = None
            current_scores = []
            in_cols, in_students, in_student_scores = False, False, False

            for prefix, event, value in parser:
                if prefix == 'cols' and event == 'start_array':
                    in_cols = True
                    continue
                if in_cols:
                    if event == 'string':
                        cols.append(value)
                    elif event == 'end_array':
                        in_cols = False
                    continue

                if prefix == 'students' and event == 'start_map':
                    in_students = True
                    if not header_written:
                        header_row = ["SBD"] + cols
                        header_cells = []
                        for val in header_row:
                            cell = openpyxl.cell.WriteOnlyCell(ws, value=val)
                            cell.font = header_font
                            cell.fill = header_fill
                            cell.alignment = header_alignment
                            cell.border = border_all
                            header_cells.append(cell)
                        ws.append(header_cells)
                        header_written = True
                    continue

                if in_students:
                    if prefix == 'students' and event == 'map_key':
                        current_sbd = value
                    elif prefix == f'students.{current_sbd}' and event == 'start_array':
                        in_student_scores = True
                        current_scores = []
                    elif in_student_scores:
                        if event in ('number', 'integer', 'double'):
                            current_scores.append(value)
                        elif event == 'null':
                            current_scores.append(None)
                        elif event == 'end_array':
                            in_student_scores = False
                            student_count += 1
                            
                            row_fill = zebra_fill if student_count % 2 == 0 else white_fill
                            row_cells = []
                            
                            sbd_cell = openpyxl.cell.WriteOnlyCell(ws, value=current_sbd)
                            sbd_cell.font = data_font
                            sbd_cell.alignment = sbd_alignment
                            sbd_cell.fill = row_fill
                            sbd_cell.border = border_all
                            row_cells.append(sbd_cell)
                            
                            for i in range(len(cols)):
                                score_val = current_scores[i] if i < len(current_scores) else None
                                score_cell = openpyxl.cell.WriteOnlyCell(ws, value=score_val)
                                score_cell.font = data_font
                                score_cell.alignment = score_alignment
                                score_cell.fill = row_fill
                                score_cell.border = border_all
                                if score_val is not None and isinstance(score_val, (int, float)):
                                    score_cell.number_format = '0.00'
                                row_cells.append(score_cell)
                            
                            ws.append(row_cells)
                            
                            if student_count % 5000 == 0:
                                loop.call_soon_threadsafe(queue.put_nowait, student_count)
                                
                    elif prefix == 'students' and event == 'end_map':
                        in_students = False

        wb.save(excel_filepath)
        wb.close()
        loop.call_soon_threadsafe(queue.put_nowait, f"DONE_{student_count}")
        
    except Exception as e:
        logger.error(f"❌ Thất bại khi phân tích tệp JSON: {str(e)}")
        loop.call_soon_threadsafe(queue.put_nowait, None)

# --- LUỒNG XỬ LÝ BACKGROUND CHO LINK OUTSIDE ---
async def process_url_background(chat_id, target_url, context: ContextTypes.DEFAULT_TYPE):
    # Trích xuất tên file từ URL để đặt tên file Excel kết quả
    parsed_filename = target_url.split("/")[-1]
    if "?" in parsed_filename:
        parsed_filename = parsed_filename.split("?")[0]
    
    base_name = os.path.splitext(parsed_filename)[0] if parsed_filename else "data"
    if not base_name.isalnum():
         base_name = "KhaChuaDiem"

    json_path = f"temp_raw_{chat_id}.json"
    excel_path = f"{base_name}_{chat_id}_converted.xlsx"
    
    status_message = await context.bot.send_message(
        chat_id=chat_id, 
        text="📥 Đã nhận liên kết! Đang tiến hành Stream tải file từ nguồn cấp dữ liệu..."
    )
    
    try:
        loop = asyncio.get_running_loop()
        
        # 1. Tải file từ link ngoài qua executor độc lập
        await loop.run_in_executor(None, download_url_stream, target_url, json_path)
        await status_message.edit_text("⏳ Đã tải xong file tạm cục bộ. Đang xử lý bóc tách cấu trúc (RAM Safe)...")
        
        queue = asyncio.Queue()
        convert_task = loop.run_in_executor(
            None, convert_heavy_json_to_excel, json_path, excel_path, queue, loop
        )
        
        while True:
            res = await queue.get()
            if res is None:
                raise Exception("Lỗi cấu trúc file JSON hoặc đường dẫn không trả về mảng dữ liệu phù hợp.")
            if isinstance(res, str) and res.startswith("DONE_"):
                total_row = res.split("_")[1]
                break
            try:
                await status_message.edit_text(f"⏳ Đang xử lý... Đã ghi thành công {res} thí sinh vào Excel.")
            except Exception:
                pass
                
        await convert_task
        await status_message.edit_text("📤 Đã chuyển đổi hoàn tất! Đang đẩy tệp Excel lên Telegram...")
        
        with open(excel_path, 'rb') as excel_file:
            await context.bot.send_document(
                chat_id=chat_id,
                document=excel_file,
                filename=f"{base_name}.xlsx",
                caption=f"🎉 Chuyển đổi thành công từ Link GitHub Raw!\n📊 Tổng số dòng dữ liệu: {total_row}\n🤖 Đã tối ưu hóa RAM an toàn tuyệt đối."
            )
            
    except Exception as e:
        logger.error(f"🚨 Sự cố khi xử lý link ngoài: {str(e)}")
        try:
            await context.bot.send_message(
                chat_id=chat_id, 
                text="💥 Không thể xử lý link này. Lý do có thể do link lỗi 404, file quá nặng, hoặc cấu trúc JSON không đúng chuẩn `cols` và `students`."
            )
        except Exception:
            pass
    finally:
        if os.path.exists(json_path): os.remove(json_path)
        if os.path.exists(excel_path): os.remove(excel_path)
        try:
            await status_message.delete()
        except Exception:
            pass

# --- XỬ LÝ SỰ KIỆN TIN NHẮN CHAT ---
async def start_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    loop = asyncio.get_running_loop()
    await loop.run_in_executor(None, clear_system_cached_files)
    msg = "👋 Bot Chuyển Đổi URL Raw JSON ➔ Excel sẵn sàng!\n\nHãy gửi trực tiếp link Raw dạng `https://raw.githubusercontent.com/...` vào đây, tôi sẽ chuyển đổi ngay lập tức."
    await update.message.reply_text(msg)

async def handle_text_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text_input = update.message.text.strip()
    
    # Biểu thức kiểm tra nếu tin nhắn gửi lên là một đường dẫn URL
    url_match = re.match(r'^(http|https)://[^\s]+$', text_input)
    
    if url_match:
        # Kích hoạt tiến trình chạy ngầm tải và phân tích từ URL
        asyncio.create_task(process_url_background(update.effective_chat.id, text_input, context))
    else:
        await update.message.reply_text("❌ Định dạng không hợp lệ. Vui lòng gửi một liên kết URL (ví dụ link raw từ GitHub của bạn).")

tg_application.add_handler(CommandHandler("start", start_command))
tg_application.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_text_message))

# --- LIFESPAN FASTAPI CONTROL ---
@asynccontextmanager
async def lifespan(app: FastAPI):
    webhook_url = f"{RENDER_EXTERNAL_URL}/telegram-webhook"
    logger.info(f"🚀 Kích hoạt Webhook External URL Stream: {webhook_url}")
    await tg_application.initialize()
    await tg_application.bot.set_webhook(url=webhook_url, drop_pending_updates=True)
    await tg_application.start()
    yield
    logger.info("🛑 Đang đóng kết nối Web Service...")
    await tg_application.bot.delete_webhook()
    await tg_application.stop()
    await tg_application.shutdown()

app = FastAPI(lifespan=lifespan)

@app.get("/")
def health_check():
    return {"status": "online", "mode": "external_url_streaming"}

@app.post("/telegram-webhook")
async def telegram_webhook(request: Request):
    try:
        req_body = await request.json()
        update = Update.de_json(req_body, tg_application.bot)
        asyncio.create_task(tg_application.process_update(update))
        return Response(status_code=status.HTTP_200_OK)
    except Exception as e:
        logger.error(f"🚨 Lỗi Webhook đầu vào: {str(e)}")
        return Response(status_code=status.HTTP_400_BAD_REQUEST)
